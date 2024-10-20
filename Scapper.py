# Import các thư viện Python cần thiết
import time
import re
import openpyxl
import os
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.edge.options import Options
from datetime import date
from selenium.webdriver.edge.service import Service
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.wait import WebDriverWait

# webdriver path
path = "G:\Learning\Code\Webdriver\msedgedriver.exe"
# Open ref.xlxs (contains district)
ref = openpyxl.load_workbook("ref.xlsx")
disLis = ref.worksheets[0]
row_count = disLis.max_row

keyword_wb = openpyxl.load_workbook("keywords.xlsx")
keyword_list = keyword_wb.worksheets[0]
keyword_row = keyword_list.max_row

# Open Results.xlsx to save the results
workBook = openpyxl.load_workbook("Results.xlsx")
workBook = openpyxl.Workbook()
workSheet = workBook.worksheets[0]
workSheet.append(
    ["District", "Name", "Longitude", "Latitude", "Type", "Rating", "Comments"]
)
workBook.save(filename="Results.xlsx")
# Setup Edge Driver configuration
service = Service(executable_path=path)
edge_options = webdriver.EdgeOptions()
edge_options.add_argument("--headless")
edge_options.add_argument("--no-sandbox")
edge_options.add_argument("--disable-dev-shm-usage")

driver = webdriver.Edge(service=service, options=edge_options)

driver.get("https://www.google.com/maps/@9.779349,105.6189045,11z?hl=vi-VN")

# Off Google Consent Form
try:
    driver.find_element(
        By.XPATH,
        "/html/body/div[3]/div[9]/div[3]/div[1]/div[1]/div[1]/div[2]/form/div[2]/div[3]/div/input[1]",
    ).click()
except:
    pass

# Search and save the result
searchBox = driver.find_element(By.ID, "searchboxinput")
for k in range(1, keyword_row + 1):
    search_key = keyword_list["A" + str(k)].value

    for i in range(1, row_count + 1):
        print(str(i) + "/" + str(row_count))
        print(search_key + " tại " + disLis["A" + str(i)].value)

        # Từ khóa tìm kiếm = {quận/huyện, tỉnh} + {đối tượng tìm kiếm}
        searchBox.send_keys(search_key + " tại " + disLis["A" + str(i)].value)
        searchBox.send_keys(Keys.ENTER)
        time.sleep(3)

        while True:
            try:
                # Pull panel to load more locations
                for j in range(0, 1):
                    # searchResults = driver.find_elements(By.XPATH, '//*[@id="QA0Szd"]//a')

                    # searchResults.location_once_scrolled_into_view
                    searchResults = driver.find_elements(
                        By.CSS_SELECTOR, "div[class^='Nv2PK']"
                    )
                    print(searchResults.get_attribute("outerHTML"))
                    if searchResults:
                        driver.execute_script(
                            "arguments[0].scrollIntoView();", searchResults[-1]
                        )
                        time.sleep(1)
                    else:
                        break

                    # print(len(searchResults))
                # Save the records
                for result in searchResults:
                    childInfo = result.find_elements(By.TAG_NAME, "a")
                    print(childInfo[0].get_attribute("outerHTML"))
                    childSpan = result.find_elements(By.XPATH, ".//span/span")

                    if len(childInfo) > 0:
                        name = childInfo[0].get_attribute("aria-label")
                        https = childInfo[0].get_attribute("href")
                        lat = re.search(r"!3d[0-9.]+", https).group()[3:]
                        long = re.search(r"!4d[0-9.]+", https).group()[3:]

                        rating = ""
                        comments = ""
                        for result2 in childSpan:
                            childRatingResults = result2.find_elements(
                                By.CLASS_NAME, "MW4etd"
                            )
                            childCommentResults = result2.find_elements(
                                By.CLASS_NAME, "UY7F9"
                            )

                            if len(childRatingResults) > 0:
                                rating = childRatingResults[0].text

                            if len(childCommentResults) > 0:
                                comments = childCommentResults[0].text
                                break

                        workSheet.append(
                            [
                                disLis["A" + str(i)].value,
                                name,
                                long,
                                lat,
                                search_key,
                                rating,
                                comments,
                            ]
                        )
                    else:
                        continue

                driver.find_element(By.XPATH, '//*[@id="sb_cb50"]').click()
                time.sleep(3)
            except:
                break

        searchBox.clear()
        workBook.save("Results.xlsx")

# Quit ChromeDriver
driver.quit()
